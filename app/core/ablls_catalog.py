import logging
import re
from pathlib import Path

import openpyxl
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.ablls_task import ABLLSTask

logger = logging.getLogger(__name__)

SECTION_NAMES = {
    "A": "Cooperation and Reinforcer Effectiveness",
    "B": "Visual Performance",
    "C": "Receptive Language",
    "D": "Motor Imitation",
    "E": "Vocal Imitation",
    "F": "Requests",
    "G": "Labeling",
    "H": "Intraverbals",
    "I": "Spontaneous Vocalizations",
    "J": "Syntax and Grammar",
    "K": "Play and Leisure",
    "L": "Social Interaction",
    "M": "Group Instruction",
    "N": "Classroom Routines",
    "P": "Generalized Responding",
    "Q": "Reading",
    "R": "Math",
    "S": "Writing",
    "T": "Spelling",
    "U": "Dressing",
    "V": "Eating",
    "W": "Grooming",
    "X": "Toileting",
    "Y": "Gross Motor",
    "Z": "Fine Motor",
}

_TASK_CODE_RE = re.compile(r"^([A-Z]{1,2})(\d+)$")
_SCORE_TOKEN_RE = re.compile(r"(\d+)\s*=")


def _extract_max_score(criteria: str) -> int:
    numbers = [int(token) for token in _SCORE_TOKEN_RE.findall(criteria or "")]
    if numbers:
        return max(numbers)
    return 1


def _sheet_to_section(sheet_name: str) -> str | None:
    normalized = (sheet_name or "").strip().upper()
    if normalized.startswith("SECTION "):
        normalized = normalized.replace("SECTION ", "", 1).strip()
    return normalized if normalized in SECTION_NAMES else None


def _row_to_task(source_sheet: str, row: tuple) -> ABLLSTask | None:
    raw_code = row[0]
    if not isinstance(raw_code, str):
        return None

    code = raw_code.strip().upper()
    match = _TASK_CODE_RE.match(code)
    if not match:
        return None

    section_code = match.group(1)
    item_number = int(match.group(2))
    objective = str(row[1]).strip() if row[1] is not None else ""
    criteria = str(row[2]).strip() if row[2] is not None else "1= да, 0= нет"
    max_score = _extract_max_score(criteria)
    section_name = SECTION_NAMES.get(section_code, f"Section {section_code}")

    return ABLLSTask(
        code=code,
        section_code=section_code,
        section_name=section_name,
        item_number=item_number,
        objective=objective,
        criteria=criteria,
        max_score=max_score,
        source_sheet=source_sheet,
    )


def load_tasks_from_workbook(workbook_path: str | Path) -> list[ABLLSTask]:
    workbook_file = Path(workbook_path)
    if not workbook_file.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_file}")

    workbook = openpyxl.load_workbook(workbook_file, data_only=True)
    tasks_by_code: dict[str, ABLLSTask] = {}

    for sheet_name in workbook.sheetnames:
        section_from_sheet = _sheet_to_section(sheet_name)
        worksheet = workbook[sheet_name]
        for row in worksheet.iter_rows(values_only=True):
            task = _row_to_task(sheet_name, row)
            if not task:
                continue
            if section_from_sheet and task.section_code != section_from_sheet:
                continue
            tasks_by_code[task.code] = task

    ordered_codes = sorted(
        tasks_by_code.keys(),
        key=lambda code: (tasks_by_code[code].section_code, tasks_by_code[code].item_number),
    )
    return [tasks_by_code[code] for code in ordered_codes]


def ensure_ablls_catalog(db: Session, workbook_path: str | Path) -> int:
    has_any = db.execute(select(ABLLSTask.code).limit(1)).scalar_one_or_none()
    if has_any:
        return 0

    try:
        tasks = load_tasks_from_workbook(workbook_path)
    except Exception:
        db.rollback()
        logger.exception("Failed to load ABLLS catalog from %s", workbook_path)
        return 0

    if not tasks:
        logger.warning("ABLLS catalog workbook parsed, but no tasks were found.")
        return 0

    db.add_all(tasks)
    db.commit()
    logger.info("ABLLS catalog loaded: %s tasks", len(tasks))
    return len(tasks)
